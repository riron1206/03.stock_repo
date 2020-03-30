#! python3
# -*- coding: utf-8 -*-
#kaisya_list.py

#IPOリストのcsvファイルを元にエクセルシートを更新する
#引数は会社データとする

#使用するモジュールのインポート
from openpyxl import load_workbook
import pandas as pd

#マスタのIPOリストを元にエクセルファイルを更新する
def master_excel(k_data):
    #管理ファイルを開き、管理シートをロードする
    wb = load_workbook(r"C:\Users\User\code\01_自動化\管理.xlsx")
    ws = wb['イベント管理']

    #マスタのcsvファイルを、リストデータとして読込む
    m_list = pd.read_csv(r"C:\Users\User\code\01_自動化\01_ipo_file\ipo_list_master.csv",encoding="shift_jis",header=None).values.tolist()
    
    #初期化が必要！！
    
    #管理シートのIPO一覧にIPO銘柄情報を書込む
    for loop in range(len(m_list)):
        ws.cell(row = k_data[2][0], column = loop + 6, value=m_list[loop][0])
        ws.cell(row = k_data[2][0] + 1, column = loop + 6, value=m_list[loop][1])
        ws.cell(row = k_data[2][0] + 2, column = loop + 6, value=m_list[loop][2])
        ws.cell(row = k_data[2][0] + 3, column = loop + 6, value=m_list[loop][3])

    #エクセルファイルを保存しクローズする
    wb.save(r"C:\Users\User\code\01_自動化\管理.xlsx")
    wb.close()



#マスタと指定された会社のIPOリストを元にエクセルファイルを更新する
def kaisya_excel(k_data):
    #管理ファイルを開き、管理シートをロードする
    wb = load_workbook(r"C:\Users\User\code\01_自動化\管理.xlsx")
    ws = wb['イベント管理']

    #マスタのcsvファイルを、リストデータとして読込む
    #各会社のcsvファイルを、リストデータとして読込む
    m_list = pd.read_csv(r"C:\Users\User\code\01_自動化\01_ipo_file\ipo_list_master.csv",encoding="shift_jis",header=None).values.tolist()
    k_list = pd.read_csv(r"C:\Users\User\code\01_自動化\01_ipo_file\ipo_list_" + str(k_data[0][0]) + ".csv",encoding="shift_jis",header=None).values.tolist()

    #初期化後、読み込んだIPOリストと、各会社の申込リストを比較し、申込可能な銘柄に印をつける
    for syokika in range(30):
        ws.cell(row = k_data[2][0], column = syokika + 6, value="")

    for m_loop in range(len(m_list)):
        for k_loop in range(len(k_list)):
            if m_list[m_loop][0] == k_list[k_loop][0]:
                ws.cell(row = k_data[2][0], column = m_loop + 6, value="●")
            else:
                pass

    #エクセルファイルを保存しクローズする
    wb.save(r"C:\Users\User\code\01_自動化\管理.xlsx")
    wb.close()




#マスタのIPOリストを元にエクセルファイルを更新する
def bunbai_excel(bun_data):
    #管理ファイルを開き、管理シートをロードする
    wb = load_workbook(r"C:\Users\User\code\01_自動化\管理.xlsx")
    ws = wb['イベント管理']

    #マスタのcsvファイルを、リストデータとして読込む
    b_list = pd.read_csv(r"C:\Users\User\code\01_自動化\03_bunbai_file\\" + str(bun_data[0][0]) + "_list.csv",encoding="shift_jis",header=None).values.tolist()
    
    #初期化が必要！！
    
    #管理シートの立会外分売一覧に銘柄情報を書込む
    for loop in range(len(b_list)):
        ws.cell(row = bun_data[2][0], column = loop + 6, value=b_list[loop][0])
        ws.cell(row = bun_data[2][0] + 1, column = loop + 6, value=b_list[loop][1])
        ws.cell(row = bun_data[2][0] + 2, column = loop + 6, value=b_list[loop][2])
        ws.cell(row = bun_data[2][0] + 3, column = loop + 6, value=b_list[loop][3])

    #エクセルファイルを保存しクローズする
    wb.save(r"C:\Users\User\code\01_自動化\管理.xlsx")
    wb.close()